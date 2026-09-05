"""Streamlit UI for the AI Audit Assistant.

Phase 6 presents the transaction and document pipelines as one audit platform:

    transaction intelligence ----\
                               -> shared review workflow -> reports / queue
    document audit ------------/

The frontend is intentionally thin. Detection, scoring and routing are computed
by the FastAPI backend. Streamlit stores completed analyses in session state so
navigation, downloads and drill-downs do not rerun the ML/audit pipelines.

Run:
    uvicorn app.main:app --reload
    streamlit run frontend/streamlit_app.py
"""
from __future__ import annotations

import io
import json
import sys
from pathlib import Path

import pandas as pd
import requests
import streamlit as st
from openpyxl.styles import Alignment, Font, PatternFill

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from config import API_BASE_URL  # noqa: E402
from app.ai.audit_chat import answer_audit_question  # noqa: E402

PROJECT_ROOT = Path(__file__).resolve().parent.parent
SCHEMA_EXAMPLES_DIR = PROJECT_ROOT / "data" / "schema_examples"
EVALUATION_DIR = PROJECT_ROOT / "data" / "evaluation"

st.set_page_config(
    page_title="AI Audit — Fraud Screening",
    page_icon="🛡️",
    layout="wide",
    initial_sidebar_state="expanded",
)

JSON_TEMPLATE = {
    "vendor": "Example Vendor",
    "invoice_number": "INV-0001",
    "date": "2026-08-10",
    "payment_date": "2026-08-20",
    "amount": 10000,
    "tax": 1800,
    "total": 11800,
    "category": "Equipment",
    "notes": "Example invoice",
}

CSV_TEMPLATE = """vendor,invoice_number,date,payment_date,amount,tax,total,category,notes
Example Vendor,INV-0001,2026-08-10,2026-08-20,10000,1800,11800,Equipment,Example invoice
"""

STATUS_ICON = {"PASS": "✅", "WARNING": "⚠️", "FAIL": "❌"}
RISK_ORDER = ["LOW", "MEDIUM", "HIGH", "CRITICAL"]


# ---------------------------------------------------------------------------
# Visual shell — inspired by the user's reference: dark navigation rail,
# generous white workspace, KPI strip, central work table and insight panel.
# ---------------------------------------------------------------------------

def inject_styles() -> None:
    st.markdown(
        """
        <style>
        :root {
            --navy-950: #061a33;
            --navy-900: #0a2342;
            --navy-800: #10335a;
            --blue: #0f6fe8;
            --cyan: #16b9d4;
            --ink: #16233a;
            --muted: #66758a;
            --line: #e5eaf1;
            --bg: #f6f8fb;
            --card: #ffffff;
            --danger: #e64855;
            --warning: #f39a3f;
            --success: #1c9b76;
        }

        .stApp {
            background: var(--bg);
            color: var(--ink);
        }
        [data-testid="stHeader"] { background: transparent; }
        [data-testid="stSidebar"] {
            background: linear-gradient(180deg, var(--navy-950) 0%, var(--navy-900) 100%);
            border-right: 1px solid rgba(255,255,255,.06);
            min-width: 250px;
        }
        [data-testid="stSidebar"] > div { padding-top: 1.15rem; }
        [data-testid="stSidebar"] [data-testid="stMarkdownContainer"],
        [data-testid="stSidebar"] label,
        [data-testid="stSidebar"] p,
        [data-testid="stSidebar"] span { color: #eef5ff; }
        [data-testid="stSidebar"] [role="radiogroup"] { gap: .22rem; }
        [data-testid="stSidebar"] [role="radiogroup"] label {
            padding: .72rem .78rem;
            border-radius: .58rem;
            margin: .04rem 0;
            transition: all .16s ease;
            font-weight: 560;
        }
        [data-testid="stSidebar"] [role="radiogroup"] label:hover {
            background: rgba(255,255,255,.08);
        }
        [data-testid="stSidebar"] [role="radiogroup"] label:has(input:checked) {
            background: linear-gradient(90deg, #0f6fe8, #0861cf);
            box-shadow: 0 6px 18px rgba(0,93,206,.28);
        }
        [data-testid="stSidebar"] hr { border-color: rgba(255,255,255,.12); }

        .block-container {
            max-width: 1520px;
            padding-top: 1.55rem;
            padding-bottom: 3rem;
            padding-left: 2rem;
            padding-right: 2rem;
        }
        h1, h2, h3, h4 { color: var(--ink); }
        .audit-eyebrow {
            color: var(--blue);
            font-size: .73rem;
            font-weight: 800;
            letter-spacing: .12em;
            text-transform: uppercase;
            margin-bottom: .3rem;
        }
        .audit-title {
            color: var(--ink);
            font-size: 2rem;
            line-height: 1.1;
            font-weight: 760;
            letter-spacing: -.025em;
            margin-bottom: .42rem;
        }
        .audit-subtitle {
            color: var(--muted);
            font-size: .96rem;
            max-width: 820px;
            line-height: 1.55;
            margin-bottom: 1.28rem;
        }
        .section-label {
            color: var(--ink);
            font-size: 1.03rem;
            font-weight: 720;
            margin-top: .35rem;
            margin-bottom: .58rem;
        }

        .kpi-grid {
            display:grid;
            grid-template-columns: repeat(4, minmax(0,1fr));
            gap: 14px;
            margin: .25rem 0 1.2rem 0;
        }
        .kpi-card {
            background: var(--card);
            border: 1px solid var(--line);
            border-radius: 13px;
            padding: 15px 16px;
            box-shadow: 0 3px 12px rgba(22,35,58,.045);
            min-height: 92px;
        }
        .kpi-label { color: var(--muted); font-size: .78rem; font-weight: 650; }
        .kpi-value { color: var(--ink); font-size: 1.55rem; font-weight: 780; margin-top: .25rem; }
        .kpi-note { color: #8793a4; font-size: .72rem; margin-top: .2rem; }

        [data-testid="stMetric"] {
            background: var(--card);
            border: 1px solid var(--line);
            border-radius: .78rem;
            padding: .85rem .95rem;
            box-shadow: 0 2px 8px rgba(22,35,58,.04);
        }
        [data-testid="stDataFrame"] {
            border: 1px solid var(--line);
            border-radius: .72rem;
            overflow: hidden;
            background: var(--card);
            box-shadow: 0 2px 10px rgba(22,35,58,.035);
        }
        [data-testid="stExpander"] {
            background: var(--card);
            border: 1px solid var(--line);
            border-radius: .72rem;
        }
        [data-testid="stAlert"] { border-radius: .72rem; }
        [data-testid="stSelectbox"] > div > div { border-radius: .62rem; }
        .insight-card {
            background: var(--card);
            border: 1px solid var(--line);
            border-radius: .78rem;
            padding: 1rem 1rem .78rem 1rem;
            margin-bottom: .8rem;
            box-shadow: 0 2px 10px rgba(22,35,58,.035);
        }
        .insight-card h4 { margin: 0 0 .45rem 0; font-size: .98rem; }
        .muted-note { color: var(--muted); font-size: .85rem; }
        .sidebar-brand {
            color: #ffffff;
            font-size: 1.25rem;
            font-weight: 800;
            letter-spacing: -.01em;
            padding: .35rem .1rem .2rem .1rem;
        }
        .sidebar-brand .mark { color: #19c0db; margin-right: .25rem; }
        .sidebar-caption {
            color: #aebdcb;
            font-size: .76rem;
            line-height: 1.35;
            margin: 0 0 1.1rem .1rem;
        }
        /* Keep Streamlit widgets visually consistent with the light workspace. */
        div.stButton > button, div.stDownloadButton > button {
            border-radius: .58rem;
            min-height: 2.55rem;
            font-weight: 650;
            background: #ffffff !important;
            color: var(--ink) !important;
            border: 1px solid #cfd8e5 !important;
            box-shadow: 0 1px 2px rgba(15,23,42,.03);
        }
        div.stButton > button p, div.stDownloadButton > button p {
            color: inherit !important;
        }
        div.stButton > button:hover, div.stDownloadButton > button:hover {
            border-color: #8fb8ef !important;
            color: var(--blue) !important;
            background: #f8fbff !important;
        }
        div.stButton > button[kind="primary"] {
            background: var(--blue) !important;
            color: #ffffff !important;
            border-color: var(--blue) !important;
        }
        div.stButton > button[kind="primary"]:hover {
            background: #0b61cf !important;
            color: #ffffff !important;
        }
        [data-testid="stExpander"] details,
        [data-testid="stExpander"] summary {
            background: #ffffff !important;
            color: var(--ink) !important;
        }
        [data-testid="stExpander"] summary p,
        [data-testid="stExpander"] summary span {
            color: var(--ink) !important;
        }
        [data-testid="stMetricValue"], [data-testid="stMetricLabel"] {
            color: var(--ink) !important;
        }
        [data-testid="stCaptionContainer"] { color: var(--muted) !important; }
        @media(max-width: 1000px) { .kpi-grid { grid-template-columns: repeat(2,1fr); } }
        </style>
        """,
        unsafe_allow_html=True,
    )


def render_kpi_cards(cards: list[tuple[str, object, str]]) -> None:
    parts = ['<div class="kpi-grid">']
    for label, value, note in cards:
        parts.append(
            f'<div class="kpi-card"><div class="kpi-label">{label}</div>'
            f'<div class="kpi-value">{value}</div><div class="kpi-note">{note}</div></div>'
        )
    parts.append('</div>')
    st.markdown(''.join(parts), unsafe_allow_html=True)


def page_header(title: str, subtitle: str, eyebrow: str = "AI AUDIT") -> None:
    st.markdown(
        f"""
        <div class="audit-eyebrow">{eyebrow}</div>
        <div class="audit-title">{title}</div>
        <div class="audit-subtitle">{subtitle}</div>
        """,
        unsafe_allow_html=True,
    )


def sidebar_navigation() -> str:
    with st.sidebar:
        st.markdown('<div class="sidebar-brand"><span class="mark">◈</span> AI AUDIT</div>', unsafe_allow_html=True)
        st.markdown(
            '<div class="sidebar-caption">Human-led fraud screening & audit intelligence</div>',
            unsafe_allow_html=True,
        )
        page = st.radio(
            "Navigation",
            [
                "Overview",
                "Transaction Analysis",
                "Document Audit",
                "Human Review Queue",
                "Audit Assistant",
                "Reports",
            ],
            label_visibility="collapsed",
        )
        st.divider()
        st.caption("Detection evidence → review priority → human decision")
        st.caption("Anomaly ≠ fraud · Risk score ≠ fraud probability")
    return page


# ---------------------------------------------------------------------------
# API helpers
# ---------------------------------------------------------------------------

def _post_file(endpoint: str, filename: str, content: bytes, timeout: int = 60):
    return requests.post(
        f"{API_BASE_URL}{endpoint}",
        files={"file": (filename, content)},
        timeout=timeout,
    )


def _show_backend_error(resp, prefix: str) -> None:
    try:
        detail = resp.json().get("detail", "Unknown backend error")
    except Exception:
        detail = resp.text
    st.error(f"{prefix}: {detail}")


def _connection_error() -> None:
    st.error(
        f"Could not reach the backend at {API_BASE_URL}. "
        "Start it with: `uvicorn app.main:app --reload`"
    )


# ---------------------------------------------------------------------------
# Reports
# ---------------------------------------------------------------------------

def create_transaction_audit_excel_report(result: dict) -> bytes:
    output = io.BytesIO()

    transactions_df = pd.DataFrame(result.get("transactions", []))
    anomalies_df = pd.DataFrame(result.get("anomalies", []))
    review_queue_df = pd.DataFrame(result.get("review_queue", []))
    cluster_df = pd.DataFrame(result.get("cluster_summary", []))
    evaluation = result.get("evaluation")
    benchmark_df = pd.DataFrame(evaluation.get("by_pattern", [])) if evaluation else pd.DataFrame()
    schema_mapping_df = pd.DataFrame(result.get("schema_mapping", []))

    coverage = result.get("data_coverage", {})
    coverage_rows = []
    for key in [
        "minimum_requirement",
        "recommended_fields",
        "optional_fields",
        "available_fields",
        "recommended_fields_present",
        "recommended_fields_missing",
        "enabled_signals",
        "skipped_signals",
    ]:
        for value in coverage.get(key, []):
            coverage_rows.append(
                {
                    "Coverage type": key.replace("_", " ").title(),
                    "Field / signal": value,
                }
            )
    if coverage.get("schema_tier"):
        coverage_rows.append(
            {"Coverage type": "Schema Tier", "Field / signal": coverage["schema_tier"]}
        )
    if coverage.get("recommended_coverage") is not None:
        coverage_rows.append(
            {
                "Coverage type": "Recommended Context Coverage",
                "Field / signal": f"{float(coverage['recommended_coverage']):.1%}",
            }
        )
    data_coverage_df = pd.DataFrame(coverage_rows)

    risk_summary = result.get("risk_summary", {})
    summary_df = pd.DataFrame(
        [
            {"Metric": "Source file", "Value": result.get("source_file", "n/a")},
            {"Metric": "Transactions analysed", "Value": result.get("row_count", 0)},
            {"Metric": "Passed automated screening", "Value": result.get("normal_count", 0)},
            {"Metric": "Detected anomalies", "Value": result.get("anomaly_count", 0)},
            {"Metric": "Anomaly rate", "Value": f"{result.get('anomaly_rate', 0):.2%}"},
            {"Metric": "Batch judgement", "Value": result.get("batch_judgement", "n/a")},
            {"Metric": "Judgement reason", "Value": result.get("batch_reason", "")},
            {"Metric": "Pending human review", "Value": risk_summary.get("review_required", 0)},
            {"Metric": "Low risk", "Value": risk_summary.get("low", 0)},
            {"Metric": "Medium risk", "Value": risk_summary.get("medium", 0)},
            {"Metric": "High risk", "Value": risk_summary.get("high", 0)},
            {"Metric": "Critical risk", "Value": risk_summary.get("critical", 0)},
            {
                "Metric": "Input schema tier",
                "Value": coverage.get("schema_tier", "n/a"),
            },
            {
                "Metric": "Recommended context coverage",
                "Value": (
                    f"{float(coverage.get('recommended_coverage')):.1%}"
                    if coverage.get("recommended_coverage") is not None
                    else "n/a"
                ),
            },
            {
                "Metric": "Skipped signals",
                "Value": ", ".join(coverage.get("skipped_signals", [])) or "None",
            },
            {
                "Metric": "Method",
                "Value": "Behavioural clustering + Isolation Forest + explicit audit patterns + human routing",
            },
        ]
    )

    metric_definitions_df = pd.DataFrame(
        [
            {"Term": "Anomaly", "Meaning": "A transaction whose behaviour is unusual enough to warrant attention. It is not proof of fraud."},
            {"Term": "Risk score", "Meaning": "Transparent 0-100 review-priority score built from triggered evidence. It is not a fraud probability."},
            {"Term": "Cluster distance", "Meaning": "Distance from the assigned behavioural cluster centre in standardized feature space."},
            {"Term": "Isolation score", "Meaning": "Isolation Forest unusualness signal; larger means easier to isolate from the ledger."},
            {"Term": "LOW", "Meaning": "No material review-priority signals; automated screening can clear the row."},
            {"Term": "MEDIUM", "Meaning": "Manager review required."},
            {"Term": "HIGH", "Meaning": "Finance Manager / Internal Audit review required."},
            {"Term": "CRITICAL", "Meaning": "Senior Audit / Fraud Investigation escalation required."},
            {"Term": "Synthetic benchmark labels", "Meaning": "Evaluation-only labels used after prediction; never model features."},
        ]
    )

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="Executive Summary", index=False)
        transactions_df.to_excel(writer, sheet_name="All Transactions", index=False)
        anomalies_df.to_excel(writer, sheet_name="Anomalies Only", index=False)
        review_queue_df.to_excel(writer, sheet_name="Human Review Queue", index=False)
        if not cluster_df.empty:
            cluster_df.to_excel(writer, sheet_name="Cluster Summary", index=False)
        if evaluation is not None:
            metrics_df = pd.DataFrame(
                [
                    {"Metric": "Precision", "Value": evaluation["precision"]},
                    {"Metric": "Recall", "Value": evaluation["recall"]},
                    {"Metric": "F1", "Value": evaluation["f1"]},
                    {"Metric": "True positives", "Value": evaluation["true_positives"]},
                    {"Metric": "False positives", "Value": evaluation["false_positives"]},
                    {"Metric": "False negatives", "Value": evaluation["false_negatives"]},
                    {"Metric": "True negatives", "Value": evaluation["true_negatives"]},
                ]
            )
            metrics_df.to_excel(writer, sheet_name="Evaluation Metrics", index=False)
            if not benchmark_df.empty:
                benchmark_df.to_excel(writer, sheet_name="Pattern Evaluation", index=False)
        metric_definitions_df.to_excel(writer, sheet_name="Metric Definitions", index=False)
        if not schema_mapping_df.empty:
            schema_mapping_df.to_excel(writer, sheet_name="Schema Mapping", index=False)
        if not data_coverage_df.empty:
            data_coverage_df.to_excel(writer, sheet_name="Data Coverage", index=False)

        workbook = writer.book
        for sheet in workbook.worksheets:
            if sheet.max_row >= 1:
                for cell in sheet[1]:
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(vertical="top")
                sheet.freeze_panes = "A2"
            for column in sheet.columns:
                letter = column[0].column_letter
                max_length = max(
                    (len(str(cell.value)) if cell.value is not None else 0 for cell in column),
                    default=0,
                )
                sheet.column_dimensions[letter].width = min(max_length + 2, 48)

        if "All Transactions" in workbook.sheetnames:
            sheet = workbook["All Transactions"]
            headers = {cell.value: cell.column for cell in sheet[1]}
            anomaly_col = headers.get("is_anomaly")
            risk_col = headers.get("risk_level")
            anomaly_fill = PatternFill(fill_type="solid", fgColor="FDE2E1")
            critical_fill = PatternFill(fill_type="solid", fgColor="F7C9C6")
            if anomaly_col:
                for row in range(2, sheet.max_row + 1):
                    if sheet.cell(row=row, column=anomaly_col).value is True:
                        for cell in sheet[row]:
                            cell.fill = anomaly_fill
                    if risk_col and sheet.cell(row=row, column=risk_col).value == "CRITICAL":
                        for cell in sheet[row]:
                            cell.fill = critical_fill

    output.seek(0)
    return output.getvalue()


def create_document_audit_excel_report(result: dict) -> bytes:
    """Create a Phase-6 document report using the same human-review vocabulary."""
    output = io.BytesIO()
    extracted = result.get("extracted", {})
    findings = result.get("findings", [])

    executive_df = pd.DataFrame(
        [
            {"Metric": "Source file", "Value": result.get("source_file", "n/a")},
            {"Metric": "Audit risk score", "Value": result.get("risk_score", 0)},
            {"Metric": "Review priority", "Value": result.get("review_priority", "n/a")},
            {"Metric": "Decision", "Value": result.get("decision", "n/a")},
            {"Metric": "Assigned reviewer", "Value": result.get("assigned_reviewer", "n/a")},
            {"Metric": "Human review required", "Value": result.get("review_required", False)},
            {"Metric": "Recommended action", "Value": result.get("recommended_action", "")},
            {"Metric": "Audit reason", "Value": result.get("audit_reason", "")},
            {"Metric": "AI explanation source", "Value": result.get("ai_summary_source", "template_fallback")},
        ]
    )

    finding_rows = []
    for finding in findings:
        evidence = "; ".join(
            f"{item.get('label')}: {item.get('value')}"
            for item in finding.get("evidence", [])
        )
        finding_rows.append(
            {
                "Finding ID": finding.get("id"),
                "Finding": finding.get("title"),
                "Status": finding.get("status"),
                "Detail": finding.get("detail"),
                "Weight": finding.get("weight"),
                "Evidence": evidence,
            }
        )
    findings_df = pd.DataFrame(finding_rows)

    extracted_df = pd.DataFrame(
        [{"Field": key, "Value": value} for key, value in extracted.items()]
    )

    queue_columns = [
        "Source Type",
        "Source File",
        "Reference",
        "Amount",
        "Risk Score",
        "Risk Level",
        "Decision",
        "Assigned Reviewer",
        "Reason",
        "Recommended Action",
    ]
    queue_rows = []
    if result.get("review_required"):
        queue_rows.append(
            {
                "Source Type": "Document",
                "Source File": result.get("source_file"),
                "Reference": extracted.get("invoice_number") or extracted.get("vendor") or "Document",
                "Amount": extracted.get("total") or extracted.get("amount"),
                "Risk Score": result.get("risk_score"),
                "Risk Level": result.get("review_priority"),
                "Decision": result.get("decision"),
                "Assigned Reviewer": result.get("assigned_reviewer"),
                "Reason": result.get("audit_reason"),
                "Recommended Action": result.get("recommended_action"),
            }
        )
    queue_df = pd.DataFrame(queue_rows, columns=queue_columns)

    definitions_df = pd.DataFrame(
        [
            {"Term": "Review priority", "Meaning": "Shared LOW/MEDIUM/HIGH/CRITICAL operational band used for both documents and transactions."},
            {"Term": "Audit risk score", "Meaning": "Additive score from non-pass deterministic/statistical findings; not a probability of fraud."},
            {"Term": "Human review", "Meaning": "A reviewer validates the evidence and retains final judgement."},
            {"Term": "AI summary", "Meaning": "Explanation generated only after the audit findings are computed; it does not decide risk."},
        ]
    )

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        executive_df.to_excel(writer, sheet_name="Executive Summary", index=False)
        findings_df.to_excel(writer, sheet_name="Findings", index=False)
        queue_df.to_excel(writer, sheet_name="Human Review Queue", index=False)
        extracted_df.to_excel(writer, sheet_name="Extracted Document", index=False)
        definitions_df.to_excel(writer, sheet_name="Metric Definitions", index=False)

        workbook = writer.book
        for sheet in workbook.worksheets:
            for cell in sheet[1]:
                cell.font = Font(bold=True)
            sheet.freeze_panes = "A2"
            for column in sheet.columns:
                letter = column[0].column_letter
                max_length = max(
                    (len(str(cell.value)) if cell.value is not None else 0 for cell in column),
                    default=0,
                )
                sheet.column_dimensions[letter].width = min(max_length + 2, 55)

        findings_sheet = workbook["Findings"]
        headers = {cell.value: cell.column for cell in findings_sheet[1]}
        status_col = headers.get("Status")
        warning_fill = PatternFill(fill_type="solid", fgColor="FFF1C7")
        fail_fill = PatternFill(fill_type="solid", fgColor="FDE2E1")
        if status_col:
            for row in range(2, findings_sheet.max_row + 1):
                status = findings_sheet.cell(row=row, column=status_col).value
                fill = fail_fill if status == "FAIL" else warning_fill if status == "WARNING" else None
                if fill:
                    for cell in findings_sheet[row]:
                        cell.fill = fill

    output.seek(0)
    return output.getvalue()


# ---------------------------------------------------------------------------
# Unified review queue helpers
# ---------------------------------------------------------------------------

def build_unified_review_queue() -> pd.DataFrame:
    rows: list[dict] = []

    tx = st.session_state.get("transaction_result")
    if tx:
        for row in tx.get("review_queue", []):
            rows.append(
                {
                    "Source": "Transaction",
                    "Reference": row.get("transaction_id", "n/a"),
                    "Subject": row.get("employee_name") or row.get("employee_id") or "n/a",
                    "Counterparty": row.get("vendor") or "n/a",
                    "Amount": row.get("amount"),
                    "Risk Score": row.get("risk_score"),
                    "Risk Level": row.get("risk_level"),
                    "Decision": row.get("decision"),
                    "Assigned Reviewer": row.get("assigned_reviewer"),
                    "Reason": row.get("audit_reason") or row.get("anomaly_reason"),
                    "Recommended Action": row.get("recommended_action"),
                }
            )

    doc = st.session_state.get("document_result")
    if doc and doc.get("review_required"):
        extracted = doc.get("extracted", {})
        rows.append(
            {
                "Source": "Document",
                "Reference": extracted.get("invoice_number") or doc.get("source_file", "Document"),
                "Subject": extracted.get("vendor") or "n/a",
                "Counterparty": extracted.get("category") or "n/a",
                "Amount": extracted.get("total") or extracted.get("amount"),
                "Risk Score": doc.get("risk_score"),
                "Risk Level": doc.get("review_priority"),
                "Decision": doc.get("decision"),
                "Assigned Reviewer": doc.get("assigned_reviewer"),
                "Reason": doc.get("audit_reason"),
                "Recommended Action": doc.get("recommended_action"),
            }
        )

    return pd.DataFrame(rows)


# ---------------------------------------------------------------------------
# Transaction result UI
# ---------------------------------------------------------------------------

@st.fragment
def render_transaction_drilldown(anomaly_df: pd.DataFrame) -> None:
    """Only this insight panel reruns when the selected transaction changes."""
    if anomaly_df.empty:
        st.info("No flagged transactions are available for drill-down.")
        return

    ids = anomaly_df["transaction_id"].astype(str).tolist()
    selected_id = st.selectbox(
        "Selected flagged transaction",
        ids,
        key="anomaly_transaction_drilldown",
    )
    row = anomaly_df.loc[
        anomaly_df["transaction_id"].astype(str).eq(selected_id)
    ].iloc[0]

    a1, a2 = st.columns(2)
    a1.metric("Risk", str(row.get("risk_level", "n/a")))
    a2.metric("Score", f"{int(row.get('risk_score', 0))}/100")

    st.markdown("**Reviewer**")
    st.write(str(row.get("assigned_reviewer", "n/a")))
    st.markdown("**Decision**")
    st.write(str(row.get("decision", "RECHECK")))
    st.markdown("**Why flagged**")
    st.write(
        str(
            row.get(
                "anomaly_reason",
                row.get("audit_reason", "Unusual transaction behaviour."),
            )
        )
    )

    active_patterns = [
        column.removeprefix("pattern_").replace("_", " ")
        for column, value in row.items()
        if column.startswith("pattern_")
        and column != "pattern_count"
        and bool(value)
    ]
    if active_patterns:
        st.markdown("**Triggered signals**")
        for pattern in active_patterns:
            st.write(f"• {pattern}")

    with st.expander("Technical evidence"):
        st.write(f"Amount: ₹{float(row.get('amount', 0)):,.2f}")
        st.write(f"Cluster: {row.get('cluster_id', 'n/a')}")
        st.write(f"Cluster distance: {float(row.get('cluster_distance', 0)):.3f}")
        st.write(f"Isolation score: {float(row.get('isolation_score', 0)):.3f}")
        st.write(f"Employee: {row.get('employee_name', row.get('employee_id', 'n/a'))}")
        if row.get("vendor") not in (None, "", "nan"):
            st.write(f"Vendor: {row.get('vendor')}")
        if row.get("category") not in (None, "", "nan"):
            st.write(f"Category: {row.get('category')}")

    st.info(str(row.get("recommended_action", "Human review if required.")))


def render_schema_coverage(result: dict) -> None:
    schema_mapping = result.get("schema_mapping", [])
    coverage = result.get("data_coverage", {})
    with st.expander("Schema mapping & analysis coverage"):
        c1, c2 = st.columns(2)
        c1.metric("Schema tier", coverage.get("schema_tier", "n/a"))
        context = coverage.get("recommended_coverage")
        c2.metric("Recommended context", f"{float(context):.0%}" if context is not None else "n/a")
        if context is not None:
            st.progress(float(context))
            st.caption("Coverage measures supplied context, not model confidence.")

        if schema_mapping:
            st.markdown("**Auto-recognized columns**")
            st.dataframe(pd.DataFrame(schema_mapping), use_container_width=True, hide_index=True)

        enabled = coverage.get("enabled_signals", [])
        skipped = coverage.get("skipped_signals", [])
        if enabled:
            st.success("Enabled: " + ", ".join(enabled))
        if skipped:
            st.warning("Skipped because supporting columns were unavailable: " + ", ".join(skipped))


def render_transaction_analysis(result: dict) -> None:
    total = result.get("row_count", 0)
    anomalies = result.get("anomaly_count", 0)
    risk_summary = result.get("risk_summary", {})
    review_count = risk_summary.get("review_required", 0)
    high_critical = risk_summary.get("high", 0) + risk_summary.get("critical", 0)

    render_kpi_cards([
        ("Total transactions", total, "Rows processed in the active ledger"),
        ("Flagged transactions", anomalies, "Suspicious behaviour requiring attention"),
        ("Pending review", review_count, "Cases routed to human reviewers"),
        ("High / critical", high_critical, "Highest-priority review workload"),
    ])

    anomaly_df = pd.DataFrame(result.get("anomalies", []))
    review_df = pd.DataFrame(result.get("review_queue", []))

    # Primary workspace: the right rail is intentionally reserved for concise
    # audit interpretation only. Charts and drill-downs sit below the review
    # queue so they do not compete with the operational table.
    left, right = st.columns([2.65, 1], gap="large")

    with left:
        st.markdown('<div class="section-label">Suspicious transaction ledger</div>', unsafe_allow_html=True)
        if anomaly_df.empty:
            st.success("No anomalous transactions were detected in this ledger.")
        else:
            preferred = [
                "transaction_id",
                "employee_name",
                "vendor",
                "category",
                "amount",
                "risk_score",
                "risk_level",
                "decision",
                "assigned_reviewer",
            ]
            cols = [c for c in preferred if c in anomaly_df.columns]
            st.dataframe(anomaly_df[cols], use_container_width=True, hide_index=True, height=390)

        st.markdown('<div class="section-label">Human review queue</div>', unsafe_allow_html=True)
        if review_df.empty:
            st.caption("No transaction currently requires human review.")
        else:
            qcols = [
                c
                for c in [
                    "transaction_id",
                    "employee_name",
                    "amount",
                    "risk_level",
                    "assigned_reviewer",
                    "recommended_action",
                ]
                if c in review_df.columns
            ]
            st.dataframe(review_df[qcols], use_container_width=True, hide_index=True, height=300)

    with right:
        st.markdown('<div class="section-label">Audit insights</div>', unsafe_allow_html=True)
        judgement = result.get("batch_judgement", "n/a")
        if judgement == "CONDITIONAL PASS":
            st.success(judgement)
        elif judgement == "SYSTEMIC REVIEW REQUIRED":
            st.error(judgement)
        else:
            st.warning(judgement)
        st.caption(result.get("batch_reason", ""))

        anomaly_rate = result.get("anomaly_rate")
        if anomaly_rate is None and total:
            anomaly_rate = anomalies / total
        if anomaly_rate is not None:
            # Some backend versions expose 0..1 while others expose percent.
            pct = float(anomaly_rate) * 100 if float(anomaly_rate) <= 1 else float(anomaly_rate)
            st.markdown(f"**Flagged share:** {pct:.1f}%")

        st.markdown(f"**Human review:** {review_count} transaction(s)")
        st.markdown(f"**High / critical:** {high_critical} transaction(s)")

        coverage = result.get("data_coverage", {})
        context = coverage.get("recommended_coverage")
        if context is not None:
            st.markdown(f"**Data context coverage:** {float(context):.0%}")

        skipped = coverage.get("skipped_signals", [])
        if skipped:
            st.caption(
                "Some context-specific checks were skipped because supporting columns were not supplied. "
                "See Schema mapping & analysis coverage below."
            )
        else:
            st.caption("All available context-specific checks were enabled for this ledger.")

    # Supporting analytics belong below the operational review table.
    st.markdown('<div class="section-label">Risk distribution</div>', unsafe_allow_html=True)
    risk_chart = pd.DataFrame(
        {
            "Risk": RISK_ORDER,
            "Transactions": [
                risk_summary.get("low", 0),
                risk_summary.get("medium", 0),
                risk_summary.get("high", 0),
                risk_summary.get("critical", 0),
            ],
        }
    )
    st.bar_chart(risk_chart.set_index("Risk"), height=260)

    st.markdown('<div class="section-label">Selected transaction investigation</div>', unsafe_allow_html=True)
    render_transaction_drilldown(anomaly_df)

    render_schema_coverage(result)

    evaluation = result.get("evaluation")
    if evaluation is not None:
        with st.expander("Synthetic benchmark evaluation"):
            st.caption(
                "Evaluation labels are read only after prediction. They are not model inputs."
            )
            e1, e2, e3 = st.columns(3)
            e1.metric("Precision", f"{evaluation['precision']:.1%}")
            e2.metric("Recall", f"{evaluation['recall']:.1%}")
            e3.metric("F1", f"{evaluation['f1']:.1%}")
            by_pattern = evaluation.get("by_pattern", [])
            if by_pattern:
                pattern_df = pd.DataFrame(by_pattern)
                if "recall" in pattern_df.columns:
                    pattern_df["recall"] = pattern_df["recall"].map(lambda x: f"{x:.1%}")
                st.dataframe(pattern_df, use_container_width=True, hide_index=True)

    with st.expander("Model / cluster interpretation"):
        st.markdown(
            """
            - **K-Means cluster** = behavioural peer group, not a fraud/safe label.
            - **Cluster distance** = how far the row sits from its cluster centre after standardization.
            - **Isolation score** = global unusualness evidence from Isolation Forest.
            - **Pattern signals** = interpretable checks such as duplicate/split payments, bursts and shifts.
            - **Risk score** = review priority, not a calibrated fraud probability.
            """
        )
        clusters = result.get("cluster_summary", [])
        if clusters:
            st.dataframe(pd.DataFrame(clusters), use_container_width=True, hide_index=True)

    st.markdown('<div class="section-label">Exports</div>', unsafe_allow_html=True)
    tx_report = create_transaction_audit_excel_report(result)
    d1, d2 = st.columns(2)
    with d1:
        st.download_button(
            "Download complete audit report",
            data=tx_report,
            file_name="AI_Transaction_Audit_Report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    with d2:
        anomaly_csv = anomaly_df.to_csv(index=False).encode("utf-8") if not anomaly_df.empty else b""
        st.download_button(
            "Download suspicious transactions CSV",
            data=anomaly_csv,
            file_name="suspicious_transactions.csv",
            mime="text/csv",
            use_container_width=True,
            disabled=anomaly_df.empty,
        )

    with st.expander("View first 20 processed transactions"):
        st.dataframe(result.get("preview", []), use_container_width=True, hide_index=True)


# ---------------------------------------------------------------------------
# Document result UI
# ---------------------------------------------------------------------------

def render_document_analysis(result: dict) -> None:
    findings = result.get("findings", [])
    non_pass = [f for f in findings if f.get("status") != "PASS"]

    render_kpi_cards([
        ("Audit risk score", f"{result.get('risk_score', 0)}/100", "Operational review-priority score"),
        ("Review priority", result.get("review_priority", "n/a"), "LOW / MEDIUM / HIGH / CRITICAL"),
        ("Non-pass findings", len(non_pass), "Warnings and failed deterministic checks"),
        ("Human review", "Required" if result.get("review_required") else "Not required", "Final judgement remains human-led"),
    ])
    left, right = st.columns([2.15, 1], gap="large")

    with left:
        st.markdown('<div class="section-label">Document audit findings</div>', unsafe_allow_html=True)
        for finding in findings:
            icon = STATUS_ICON.get(finding.get("status"), "•")
            with st.expander(
                f"{icon} {finding.get('title', 'Finding')} — {finding.get('status', 'n/a')}",
                expanded=finding.get("status") != "PASS",
            ):
                st.write(finding.get("detail", ""))
                evidence = finding.get("evidence", [])
                if evidence:
                    st.markdown("**Evidence**")
                    for item in evidence:
                        st.write(f"• {item.get('label')}: {item.get('value')}")

        with st.expander("Extracted structured data"):
            st.json(result.get("extracted", {}))

    with right:
        st.markdown('<div class="section-label">Review routing</div>', unsafe_allow_html=True)
        priority = result.get("review_priority", "n/a")
        if priority == "LOW":
            st.success(priority)
        elif priority == "MEDIUM":
            st.warning(priority)
        else:
            st.error(priority)
        st.markdown("**Decision**")
        st.write(result.get("decision", "n/a"))
        st.markdown("**Assigned reviewer**")
        st.write(result.get("assigned_reviewer", "n/a"))
        st.markdown("**Recommended action**")
        st.info(result.get("recommended_action", ""))
        st.caption(result.get("audit_reason", ""))

        st.markdown('<div class="section-label">Audit explanation</div>', unsafe_allow_html=True)
        source = result.get("ai_summary_source", "template_fallback")
        st.caption(
            "LLM explanation from precomputed findings only."
            if source == "llm"
            else "Deterministic template explanation (no LLM key or fallback path)."
        )
        st.markdown(result.get("ai_summary", "No summary available."))

    st.divider()
    st.caption(
        "Phase 6 uses the same operational LOW/MEDIUM/HIGH/CRITICAL review hierarchy for "
        "documents and transaction ledgers. Humans retain final judgement."
    )

    report = create_document_audit_excel_report(result)
    st.download_button(
        "Download document audit report",
        data=report,
        file_name="AI_Document_Audit_Report.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ---------------------------------------------------------------------------
# Pages
# ---------------------------------------------------------------------------

def render_overview_page() -> None:
    page_header(
        "Fraud Screening — Overview",
        "One workspace for transaction intelligence, document controls, human review and audit reporting.",
    )

    tx = st.session_state.get("transaction_result")
    doc = st.session_state.get("document_result")
    queue = build_unified_review_queue()

    total_tx = tx.get("row_count", 0) if tx else 0
    flags = tx.get("anomaly_count", 0) if tx else 0
    pending = len(queue)
    critical = 0
    if tx:
        critical += tx.get("risk_summary", {}).get("critical", 0)
    if doc and doc.get("review_priority") == "CRITICAL" and doc.get("review_required"):
        critical += 1

    render_kpi_cards([
        ("Transactions analysed", total_tx, "Current cached transaction analysis"),
        ("Suspicious flags", flags, "Transactions flagged by combined evidence"),
        ("Pending reviews", pending, "Unified transaction + document queue"),
        ("Critical escalations", critical, "Cases requiring senior investigation"),
    ])
    left, right = st.columns([2.35, 1], gap="large")
    with left:
        st.markdown('<div class="section-label">Current review workload</div>', unsafe_allow_html=True)
        if queue.empty:
            st.info(
                "No review items are currently cached. Run a transaction analysis or document audit from the sidebar."
            )
        else:
            shown = queue[
                [
                    c
                    for c in [
                        "Source",
                        "Reference",
                        "Subject",
                        "Amount",
                        "Risk Level",
                        "Assigned Reviewer",
                    ]
                    if c in queue.columns
                ]
            ].head(12)
            st.dataframe(shown, use_container_width=True, hide_index=True, height=400)

        if tx:
            st.markdown('<div class="section-label">Transaction risk distribution</div>', unsafe_allow_html=True)
            rs = tx.get("risk_summary", {})
            chart = pd.DataFrame(
                {
                    "Risk": RISK_ORDER,
                    "Transactions": [rs.get("low", 0), rs.get("medium", 0), rs.get("high", 0), rs.get("critical", 0)],
                }
            )
            st.bar_chart(chart.set_index("Risk"), height=260)

    with right:
        st.markdown('<div class="section-label">Reports & insights</div>', unsafe_allow_html=True)
        if tx:
            st.markdown("**Transaction batch**")
            judgement = tx.get("batch_judgement", "n/a")
            if judgement == "CONDITIONAL PASS":
                st.success(judgement)
            elif judgement == "SYSTEMIC REVIEW REQUIRED":
                st.error(judgement)
            else:
                st.warning(judgement)
            st.caption(tx.get("batch_reason", ""))
            coverage = tx.get("data_coverage", {})
            if coverage.get("recommended_coverage") is not None:
                st.metric("Data context coverage", f"{float(coverage['recommended_coverage']):.0%}")

        if doc:
            st.markdown("**Latest document audit**")
            st.metric("Document priority", doc.get("review_priority", "n/a"))
            st.write(doc.get("decision", "n/a"))
            st.caption(doc.get("source_file", ""))

        if not tx and not doc:
            st.markdown(
                """
                **Start with either workflow:**
                - Transaction Analysis for CSV/XLSX ledgers
                - Document Audit for invoice/expense evidence

                Both feed the same human-review hierarchy.
                """
            )


def render_transaction_input() -> None:
    st.markdown('<div class="section-label">Input requirements</div>', unsafe_allow_html=True)
    st.markdown(
        """
| Level | Fields | What happens |
|---|---|---|
| **Minimum — required** | `transaction_id`, `date`, `amount`, plus `employee_id` **or** `employee_name` | ML screening can run. |
| **Recommended context** | `vendor`, `category`, `location`, `department`, `manager_name`, `payment_method` | Enables stronger context-specific signals and routing. |
| **Optional** | `description` and unrelated extra columns | Preserved where useful; not required by the detector. |
"""
    )
    st.caption(
        "Auto-recognition handles common aliases and small spelling mistakes. Missing recommended fields disable only dependent checks."
    )

    with st.expander("Templates and prepared demonstrations", expanded=True):
        min_path = SCHEMA_EXAMPLES_DIR / "01_minimum_required_sample.xlsx"
        med_path = SCHEMA_EXAMPLES_DIR / "02_medium_context_typo_sample.xlsx"
        t1, t2 = st.columns(2)
        with t1:
            if min_path.exists():
                st.download_button(
                    "Download minimum Excel example",
                    data=min_path.read_bytes(),
                    file_name=min_path.name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
            st.caption("Minimum columns accepted by the ML pipeline.")
        with t2:
            if med_path.exists():
                st.download_button(
                    "Download typo-recognition example",
                    data=med_path.read_bytes(),
                    file_name=med_path.name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
            st.caption("Richer context with intentional spelling mistakes in headers.")

        demos = {
            "Minimum required schema": min_path,
            "Medium context + spelling mistakes": med_path,
            "Mixed fraud-pattern benchmark": EVALUATION_DIR / "12_mixed_all_patterns.xlsx",
            "Systemic >50% anomaly benchmark": EVALUATION_DIR / "21_systemic_high_anomaly_60.xlsx",
        }
        choice = st.selectbox("Prepared demo", list(demos), key="prepared_tx_demo")
        if st.button("Run prepared transaction demo", use_container_width=True):
            path = demos[choice]
            if not path.exists():
                st.error(f"Prepared demo file is missing: {path.name}")
            else:
                try:
                    with st.spinner("Analysing prepared ledger..."):
                        resp = _post_file("/transactions/analyze", path.name, path.read_bytes())
                    if resp.status_code == 200:
                        st.session_state["transaction_result"] = resp.json()
                        st.session_state["transaction_result_filename"] = path.name
                        st.rerun()
                    else:
                        _show_backend_error(resp, "Prepared demo failed")
                except requests.exceptions.ConnectionError:
                    _connection_error()

    st.markdown('<div class="section-label">Upload transaction ledger</div>', unsafe_allow_html=True)
    uploaded = st.file_uploader(
        "CSV or XLSX ledger",
        type=["csv", "xlsx"],
        key="transaction_ledger",
    )
    if uploaded is not None:
        st.caption(f"Selected: {uploaded.name}")
        if st.button("Analyse transaction ledger", type="primary", use_container_width=True):
            try:
                with st.spinner("Validating schema, engineering features and screening transactions..."):
                    resp = _post_file(
                        "/transactions/analyze",
                        uploaded.name,
                        uploaded.getvalue(),
                    )
                if resp.status_code == 200:
                    st.session_state["transaction_result"] = resp.json()
                    st.session_state["transaction_result_filename"] = uploaded.name
                    st.rerun()
                else:
                    _show_backend_error(resp, "Analysis could not start")
            except requests.exceptions.ConnectionError:
                _connection_error()


def render_transaction_page() -> None:
    page_header(
        "Transaction Fraud Screening",
        "Upload a ledger, auto-map its schema, detect unusual behaviour and route suspicious rows for human review.",
    )
    cached = st.session_state.get("transaction_result")
    if cached is None:
        render_transaction_input()
        return

    h1, h2 = st.columns([5, 1])
    with h1:
        filename = st.session_state.get("transaction_result_filename", cached.get("source_file", "ledger"))
        st.caption(f"Showing persisted analysis for **{filename}**. Drill-downs and downloads do not rerun the backend pipeline.")
    with h2:
        if st.button("New analysis", use_container_width=True):
            for key in ["transaction_result", "transaction_result_filename", "anomaly_transaction_drilldown"]:
                st.session_state.pop(key, None)
            st.rerun()
    render_transaction_analysis(cached)


def render_document_input() -> None:
    st.markdown('<div class="section-label">Document input</div>', unsafe_allow_html=True)
    st.info(
        "For the most predictable extraction, use the JSON/CSV template. PDF and image files remain supported for the demo path."
    )

    d1, d2 = st.columns(2)
    with d1:
        st.download_button(
            "Download JSON template",
            data=json.dumps(JSON_TEMPLATE, indent=4),
            file_name="invoice_template.json",
            mime="application/json",
            use_container_width=True,
        )
    with d2:
        st.download_button(
            "Download CSV template",
            data=CSV_TEMPLATE,
            file_name="invoice_template.csv",
            mime="text/csv",
            use_container_width=True,
        )

    with st.expander("Try a bundled document sample"):
        try:
            samples_resp = requests.get(f"{API_BASE_URL}/audit/samples", timeout=10)
            samples = samples_resp.json() if samples_resp.status_code == 200 else []
        except requests.exceptions.ConnectionError:
            samples = []
        if samples:
            sample = st.selectbox("Sample document", samples, key="document_sample_choice")
            if st.button("Run sample document audit", use_container_width=True):
                try:
                    with st.spinner("Running document audit..."):
                        resp = requests.get(f"{API_BASE_URL}/audit/sample/{sample}", timeout=30)
                    if resp.status_code == 200:
                        st.session_state["document_result"] = resp.json()
                        st.session_state["document_result_filename"] = sample
                        st.rerun()
                    else:
                        _show_backend_error(resp, "Sample audit failed")
                except requests.exceptions.ConnectionError:
                    _connection_error()
        else:
            st.caption("Start the backend to load bundled samples.")

    uploaded = st.file_uploader(
        "Invoice or expense document",
        type=["pdf", "csv", "json", "png", "jpg", "jpeg"],
        key="document_uploader",
    )
    if uploaded is not None:
        ext = uploaded.name.lower().rsplit(".", 1)[-1]
        if ext in {"png", "jpg", "jpeg"}:
            st.warning("Image input uses OCR and may require local Tesseract installation.")
        elif ext == "pdf":
            st.caption("PDF text extraction is used; structured JSON/CSV remains more predictable.")

        if st.button("Run document audit", type="primary", use_container_width=True):
            try:
                with st.spinner("Extracting, applying audit controls and routing review..."):
                    resp = _post_file("/audit/upload", uploaded.name, uploaded.getvalue(), timeout=30)
                if resp.status_code == 200:
                    st.session_state["document_result"] = resp.json()
                    st.session_state["document_result_filename"] = uploaded.name
                    st.rerun()
                else:
                    _show_backend_error(resp, "Document audit failed")
            except requests.exceptions.ConnectionError:
                _connection_error()


def render_document_page() -> None:
    page_header(
        "Document Audit",
        "Extract invoice evidence, run deterministic controls and feed findings into the same human-review hierarchy as transactions.",
    )
    cached = st.session_state.get("document_result")
    if cached is None:
        render_document_input()
        return

    h1, h2 = st.columns([5, 1])
    with h1:
        filename = st.session_state.get("document_result_filename", cached.get("source_file", "document"))
        st.caption(f"Showing persisted audit for **{filename}**. Navigation and report downloads do not rerun extraction or scoring.")
    with h2:
        if st.button("New document", use_container_width=True):
            st.session_state.pop("document_result", None)
            st.session_state.pop("document_result_filename", None)
            st.rerun()
    render_document_analysis(cached)


def render_review_queue_page() -> None:
    page_header(
        "Human Review Queue",
        "Unified operational queue for suspicious transactions and document findings that require human judgement.",
    )
    queue = build_unified_review_queue()
    if queue.empty:
        st.info("No cached analysis currently requires human review.")
        return

    r1, r2, r3, r4 = st.columns(4)
    r1.metric("Pending items", len(queue))
    r2.metric("Transaction items", int((queue["Source"] == "Transaction").sum()))
    r3.metric("Document items", int((queue["Source"] == "Document").sum()))
    r4.metric("High / critical", int(queue["Risk Level"].isin(["HIGH", "CRITICAL"]).sum()))

    st.markdown('<div class="section-label">Review workload</div>', unsafe_allow_html=True)
    st.dataframe(queue, use_container_width=True, hide_index=True, height=520)
    st.caption(
        "This queue is a routing aid, not an automated fraud verdict. Reviewers validate supporting evidence and make the final decision."
    )



def render_audit_assistant_page() -> None:
    """Phase 7 grounded conversational interface.

    It queries only results already persisted in Streamlit session state. No
    upload is reprocessed and no risk/anomaly decision is recomputed here.
    """
    page_header(
        "Audit Assistant",
        "Ask questions about the currently computed transaction and document results. Answers are retrieved from audit evidence; the assistant does not decide whether fraud occurred.",
    )

    tx = st.session_state.get("transaction_result")
    doc = st.session_state.get("document_result")

    scope_parts = []
    if tx:
        scope_parts.append(f"Transaction ledger: {tx.get('source_file', 'cached analysis')}")
    if doc:
        scope_parts.append(f"Document: {doc.get('source_file', 'cached audit')}")

    if not scope_parts:
        st.info(
            "No audit result is cached yet. Run **Transaction Analysis** or **Document Audit** first; "
            "the assistant intentionally has nothing to answer from until evidence exists."
        )
        return

    top_left, top_right = st.columns([5, 1])
    with top_left:
        st.caption("Grounded in: **" + " · ".join(scope_parts) + "**")
    with top_right:
        if st.button("Clear chat", use_container_width=True):
            st.session_state["audit_chat_history"] = []
            st.rerun()

    st.markdown('<div class="section-label">Suggested questions</div>', unsafe_allow_html=True)
    suggestions = []
    if tx:
        anomalies = tx.get("anomalies", []) or []
        if anomalies:
            example_id = anomalies[0].get("transaction_id")
            if example_id:
                suggestions.append(f"Why was {example_id} flagged?")
        suggestions.extend([
            "Why did this batch get its current judgement?",
            "Which department has the highest anomaly rate?",
            "What signals were skipped because of schema coverage?",
            "How many pending reviews are there?",
        ])
    if doc:
        suggestions.append("Explain the current document audit")

    # Buttons are intentionally starters only; the actual conversation stays
    # in st.session_state and does not rerun the ML/audit pipeline.
    chosen_prompt = None
    if suggestions:
        cols = st.columns(min(3, len(suggestions)))
        for i, suggestion in enumerate(suggestions[:6]):
            with cols[i % len(cols)]:
                if st.button(suggestion, key=f"audit_chat_suggestion_{i}", use_container_width=True):
                    chosen_prompt = suggestion

    history = st.session_state.setdefault("audit_chat_history", [])
    if not history:
        with st.chat_message("assistant"):
            st.markdown(
                "I can explain **existing audit results**: transaction flags, employee summaries, batch judgement, "
                "risk routing, schema coverage, department anomaly rates, pending reviews and document findings. "
                "I will not turn an anomaly score into a fraud verdict."
            )

    for message in history:
        with st.chat_message(message["role"]):
            st.markdown(message["content"])
            if message["role"] == "assistant" and message.get("source"):
                source_label = (
                    "Grounded evidence + optional LLM wording"
                    if message["source"] == "llm_grounded"
                    else "Grounded deterministic answer"
                )
                st.caption(source_label)
                if message.get("facts"):
                    with st.expander("Retrieved evidence used for this answer"):
                        st.json(message["facts"])

    typed_prompt = st.chat_input("Ask about the cached audit evidence…")
    prompt = typed_prompt or chosen_prompt
    if prompt:
        history.append({"role": "user", "content": prompt})
        response = answer_audit_question(
            prompt,
            transaction_result=tx,
            document_result=doc,
            allow_llm_wording=True,
        )
        history.append(
            {
                "role": "assistant",
                "content": response["answer"],
                "source": response["source"],
                "intent": response["intent"],
                "facts": response.get("facts", {}),
            }
        )
        st.session_state["audit_chat_history"] = history
        st.rerun()

    with st.expander("What the assistant can and cannot do"):
        st.markdown(
            """
            **Can**
            - Explain why a computed transaction was flagged and how it was routed.
            - Summarize an employee's flagged transactions in the current ledger.
            - Explain batch judgement, risk levels, schema coverage and skipped signals.
            - Summarize the unified human-review workload and current document findings.

            **Cannot**
            - Recompute anomaly detection or change a risk score.
            - Declare a transaction/person/document fraudulent or legitimate.
            - Answer from external market/news data or from files that have not been analysed.
            - Treat synthetic benchmark labels as model evidence.
            """
        )

def render_reports_page() -> None:
    page_header(
        "Reports",
        "Download audit-ready workbooks from the currently persisted analyses.",
    )
    tx = st.session_state.get("transaction_result")
    doc = st.session_state.get("document_result")

    if not tx and not doc:
        st.info("Run a transaction analysis or document audit before generating reports.")
        return

    c1, c2 = st.columns(2, gap="large")
    with c1:
        st.markdown("### Transaction audit report")
        if tx:
            st.write(tx.get("source_file", "Transaction ledger"))
            st.caption(
                "Executive Summary · All Transactions · Anomalies · Human Review Queue · Clusters · Evaluation · Schema Mapping · Data Coverage"
            )
            st.download_button(
                "Download transaction workbook",
                data=create_transaction_audit_excel_report(tx),
                file_name="AI_Transaction_Audit_Report.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        else:
            st.caption("No transaction analysis is currently cached.")

    with c2:
        st.markdown("### Document audit report")
        if doc:
            st.write(doc.get("source_file", "Document"))
            st.caption(
                "Executive Summary · Findings · Human Review Queue · Extracted Document · Metric Definitions"
            )
            st.download_button(
                "Download document workbook",
                data=create_document_audit_excel_report(doc),
                file_name="AI_Document_Audit_Report.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        else:
            st.caption("No document audit is currently cached.")

    st.divider()
    st.markdown("### Reporting principle")
    st.write(
        "Both reports expose the evidence, review priority, assigned reviewer and recommended action. "
        "They do not present an anomaly or risk score as proof of fraud."
    )


def main() -> None:
    inject_styles()
    page = sidebar_navigation()

    if page == "Overview":
        render_overview_page()
    elif page == "Transaction Analysis":
        render_transaction_page()
    elif page == "Document Audit":
        render_document_page()
    elif page == "Human Review Queue":
        render_review_queue_page()
    elif page == "Audit Assistant":
        render_audit_assistant_page()
    elif page == "Reports":
        render_reports_page()


if __name__ == "__main__":
    main()
